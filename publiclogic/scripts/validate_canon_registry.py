#!/usr/bin/env python3
"""
validate_canon_registry.py  --check | --write  ROOT  WORKBOOK

The workbook must not assert what the code says. It must be told.

  --check   compare workbook against source; exit 1 on any mismatch   (CI)
  --write   regenerate the workbook's source columns from code        (local, then commit)

The source is read by EXECUTING the runtime (canon-extract.mjs imports the
modules); this script only compares. It owns, so none can be edited around:

  F  Count in source        I  Expected members (source)
  named-range span          Control!B55 (REVIEW_THRESHOLD)

A registered code-owned vocabulary with no source is a FAILURE, not a skip.
Silent additions are the bypass this exists to prevent.

Failure classes:
  DRIFT        workbook members differ from source (ordered, case-sensitive)
  COUNT        the count-in-source literal was edited (members still match)
  RANGE        the named range no longer spans the source's member count
  CONSTANT     REVIEW_THRESHOLD differs from source
  UNMAPPED     a registry vocabulary has no source at all — map it or remove it
  UNEXPORTED   a registry vocabulary claims a code const that is not exported
  UNVERIFIABLE claimed code-owned, absent from this repo (owner must resolve)
"""
import json
import re
import subprocess
import sys
import pathlib

from openpyxl import load_workbook

MODE = sys.argv[1] if len(sys.argv) > 1 else "--check"
ROOT = pathlib.Path(sys.argv[2] if len(sys.argv) > 2 else "packages/golden-path")
BOOK = pathlib.Path(sys.argv[3] if len(sys.argv) > 3 else "canon/PublicLogic_PuddleJumper_Runtime.xlsx")
HERE = pathlib.Path(__file__).parent

# Vocabularies the WORKBOOK is the source of truth for — code does not enforce
# them, so the gate does not compare them against source. Per the workbook's own
# rule (Control!A58): code is canon for vocabularies enforced at runtime; the
# workbook is canon for governance vocabularies code does not enforce.
#
# Sensitivity is workbook-owned: nothing in golden-path reads, validates, or
# rejects a sensitivity_level. It came from a superseded LogicOS SQL schema and
# was mis-registered as code-owned. Reclassifying it applies the A58 rule — it is
# not removing a check (the workbook's own Validation gate still enforces it on
# every artifact); it stops it claiming a source it never had.
WORKBOOK_OWNED = {
    "Lifecycle", "Artifact Family", "Event Type", "Nature", "Significance",
    "Status", "Risk", "Claim Status", "Object Type", "Sensitivity",
}

# Registered code-owned, but the owning source is not in this repo. The mechanism
# stays for the general case (a future row claiming a source that isn't here); it
# fails BY DESIGN. Do not "fix" it by deleting the row.
UNVERIFIABLE = {}

# --- read the source, by executing the runtime -----------------------------------
proc = subprocess.run(
    ["node", str(HERE / "canon-extract.mjs"), str(ROOT)],
    capture_output=True, text=True,
)
if proc.returncode != 0:
    print("canon-extract.mjs failed:\n" + proc.stderr, file=sys.stderr)
    sys.exit(2)
canon = json.loads(proc.stdout)
src = canon["vocabularies"]

wb = load_workbook(BOOK)
cr = wb["Canon Registry"]
problems, wrote = [], 0

# Registry rows begin at 5 and run until the "DERIVED VIEWS" sentinel or a blank
# name — scanning dynamically so an appended rogue row is still seen (UNMAPPED).
def registry_rows():
    r = 5
    while r < 200:
        name = cr.cell(r, 1).value
        if name is None:
            break
        if isinstance(name, str) and name.startswith("DERIVED VIEWS"):
            break
        yield r, name
        r += 1

for r, name in registry_rows():
    if name in UNVERIFIABLE:
        problems.append(f"UNVERIFIABLE  {name}: {UNVERIFIABLE[name]}")
        continue
    if name in WORKBOOK_OWNED:
        continue
    if name not in src:
        problems.append(f"UNMAPPED      {name}: in the registry with no source. Map it or remove it.")
        continue
    entry = src[name]
    if entry.get("unexported"):
        problems.append(f"UNEXPORTED    {name}: {entry['from']} is not exported — the extractor cannot read it "
                        f"by execution. Export it, or reclassify the row.")
        continue

    ismembers = entry["members"]

    # RANGE — the named range must span exactly the source's member count.
    rng = cr.cell(r, 2).value
    ref = wb.defined_names[rng].attr_text if rng in wb.defined_names else ""
    m = re.search(r"\$[A-Z]+\$(\d+):\$[A-Z]+\$(\d+)", ref)
    if m:
        span = int(m.group(2)) - int(m.group(1)) + 1
        if span != len(ismembers):
            problems.append(f"RANGE         {name}: {rng} spans {span} cells, source has {len(ismembers)}.")

    if MODE == "--write":
        cr.cell(r, 6).value = len(ismembers)
        cr.cell(r, 9).value = "; ".join(ismembers)
        wrote += 1
        continue

    have = [s.strip() for s in str(cr.cell(r, 9).value or "").split(";") if s.strip()]

    # DRIFT — the workbook's members vs source (ordered, case-sensitive).
    if have != ismembers:
        problems.append(
            f"DRIFT         {name}\n              source ({entry['from']}): {'; '.join(ismembers)}\n"
            f"              workbook: {'; '.join(have)}"
        )
    # COUNT — the F literal vs the row's OWN member list. Orthogonal to DRIFT and
    # NOT suppressed by it: a row can be both drifted and mis-counted, and both
    # must surface (masking one behind the other would let a partial --write pass).
    if cr.cell(r, 6).value != len(have):
        problems.append(f"COUNT         {name}: count cell says {cr.cell(r, 6).value}, the row lists {len(have)} members.")

# CONSTANT — REVIEW_THRESHOLD, held at Control!B55.
thr = canon["constants"]["REVIEW_THRESHOLD"]
if MODE == "--write":
    wb["Control"]["B55"] = thr
elif wb["Control"]["B55"].value != thr:
    problems.append(f"CONSTANT      REVIEW_THRESHOLD: workbook {wb['Control']['B55'].value}, source {thr}.")

if MODE == "--write":
    wb.save(BOOK)
    print(f"rewrote {wrote} vocabularies + REVIEW_THRESHOLD from source -> {BOOK.name}")
    if problems:
        print("\nstill unresolved (write does not touch named ranges or UNVERIFIABLE rows):\n"
              + "\n".join("  " + p for p in problems), file=sys.stderr)
        sys.exit(1)
    sys.exit(0)

if problems:
    print(f"\ncanon check FAILED - {len(problems)} problem(s)\n", file=sys.stderr)
    print("\n".join("  " + p for p in problems) + "\n", file=sys.stderr)
    print("  Fix the source, or run: python3 scripts/validate_canon_registry.py --write ...\n", file=sys.stderr)
    sys.exit(1)

n = sum(1 for _, nm in registry_rows() if nm in src and not src[nm].get("unexported"))
print(f"canon check passed - {n} code-owned vocabularies match source.")
