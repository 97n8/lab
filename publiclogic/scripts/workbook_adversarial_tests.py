import subprocess, shutil, json, sys
from openpyxl import load_workbook
from datetime import date

SRC = '/mnt/user-data/outputs/PublicLogic_PuddleJumper_Runtime_v3.3_CANON.xlsx'
RECALC = '/mnt/skills/public/xlsx/scripts/recalc.py'

# Validation row -> check name
def read(path):
    w = load_workbook(path, data_only=True)
    v = w['Validation']
    checks = {}
    for i in range(5, 36):
        n = v.cell(i, 1).value
        if n: checks[n] = (v.cell(i, 2).value, v.cell(i, 4).value)
    return w, checks

def run(name, mutate):
    p = f'/tmp/t_{name}.xlsx'
    shutil.copy(SRC, p)
    wb = load_workbook(p)
    mutate(wb)
    wb.save(p)
    r = subprocess.run(['python3', RECALC, p, '60'], capture_output=True, text=True)
    if '"status": "success"' not in r.stdout:
        return None, f'RECALC FAIL: {r.stdout[:120]}'
    return read(p)[1], None

TESTS = []
def t(name, gate, mutate, expect='CHECK'):
    TESTS.append((name, gate, mutate, expect))

def blank(sheet, cell):
    return lambda wb: wb[sheet].__setitem__(cell, None)
def setv(sheet, cell, val):
    return lambda wb: wb[sheet].__setitem__(cell, val)

t('baseline', None, lambda wb: None, None)
t('missing_family',     'Missing Family',              blank('Artifact Registry','E5'))
t('missing_nature',     'Missing Nature',              blank('Artifact Registry','F5'))
t('missing_lifecycle',  'Missing Lifecycle',           blank('Artifact Registry','G5'))
t('missing_authority',  'Missing Authority',           blank('Artifact Registry','H5'))
t('missing_claim',      'Artifacts Missing Claim Status',   blank('Artifact Registry','Q5'))
t('missing_sensitivity','Artifacts Missing Sensitivity',    blank('Artifact Registry','R5'))
t('conditional_sealed', 'CONDITIONAL Claims Sealed to ARCHIEVE', setv('Artifact Registry','G9','ARCHIEVE'))
t('orphan_event',       'Orphan Event Artifact IDs',   setv('PJ Event Log','D5','PL-DOES-NOT-EXIST-0000-00000'))
t('bogus_event_type',   'Invalid Event Log Rows',      setv('PJ Event Log','C5','BogusEventType'))
t('missing_prr_kind',   'Events Missing PRR Kind',     blank('PJ Event Log','R5'))
t('below_thresh_placed','Signals Below Threshold Not Held', setv('Signal Intake','L7','append'))
t('held_above_thresh',  'Held Signals At Or Above Threshold', setv('Signal Intake','L5','needs_review'))
t('unpreserved_signal', 'Unpreserved Signals',         setv('Signal Intake','N7','No'))
t('no_provenance',      'Signals Missing Provenance',  blank('Signal Intake','C5'))
t('invalid_signal_row', 'Invalid Signal Rows',         setv('Signal Intake','F5','not_a_kind'))
t('unregistered_form',  'Unregistered Case FORM',      setv('Intake','B12','PL-BOGUS-0000-00000'))
t('decision_no_auth',   'Decisions Missing Authority', blank('Decisions','F6'))
t('drift_append_lane',  'Vocabulary Spill (values below named range)', setv('Control','B18','PRODUCT'))
t('drift_rename_lane',  'Vocabulary Member Mismatch vs Code', setv('Control','B14','STAYY'))
t('drift_reorder_lane', 'Vocabulary Member Mismatch vs Code', lambda wb: (wb['Control'].__setitem__('B14','MUNI'), wb['Control'].__setitem__('B15','STAY')))
t('drift_case_lane',    'Vocabulary Member Mismatch vs Code', setv('Control','B14','stay'))
t('drift_rollsup',      'Vocabulary Drift vs Code',    setv('Control','B14','STAYY'))
t('drift_verb_rename',  'Vocabulary Member Mismatch vs Code', setv('Control','W14','open_casespace'))
t('drift_prr_kind',     'Vocabulary Member Mismatch vs Code', setv('Control','V16','DECISIONS'))

# negative controls: gates that must go quiet when the problem is fixed
def fix_evidence(wb):
    wb['Documents']['E10'] = date(2026, 12, 31)
    wb['Documents']['E10'].number_format = 'yyyy-mm-dd'
t('evidence_fixed', 'Overdue Evidence Requests', fix_evidence, 'OK')

def resolve_criticals(wb):
    for r in (7, 10):   # the two Critical events
        wb['PJ Event Log'].cell(r, 16).value = 'Yes'
t('criticals_resolved', 'Unresolved Critical Events', resolve_criticals, 'OK')

results = []
for name, gate, mutate, expect in TESTS:
    checks, err = run(name, mutate)
    if err:
        results.append((name, gate, 'ERROR', err)); continue
    if gate is None:
        results.append((name, '-', 'BASELINE', json.dumps({k: v[0] for k, v in checks.items() if v[1] == 'CHECK'})))
        continue
    val, sig = checks.get(gate, (None, None))
    ok = (sig == expect)
    results.append((name, gate, 'PASS' if ok else 'FAIL', f'value={val} signal={sig} expected={expect}'))

print(f'\n{"TEST":24} {"GATE":40} {"RESULT":9} DETAIL')
print('-' * 130)
for r in results:
    print(f'{r[0]:24} {str(r[1])[:39]:40} {r[2]:9} {r[3]}')
fails = [r for r in results if r[2] in ('FAIL', 'ERROR')]
print(f'\n{len(results)-1-len(fails)}/{len(results)-1} gates behaved correctly. {len(fails)} problems.')
